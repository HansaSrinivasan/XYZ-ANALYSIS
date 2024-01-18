import openpyxl
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt

# Find the first row
def findRow(sheet_obj):
    r = 1
    while(sheet_obj.cell(row=r, column=1).value == None):
        r += 1
    return r + 1

# To find the sum of each row
def annualSum(sheet_obj, r):
    col = 2
    while(str(sheet_obj.cell(row=r, column=col).value).isdigit() == False):
        col += 1
    total_sum = 0
    while(sheet_obj.cell(row=r, column=col).value != None):
        total_sum += sheet_obj.cell(row=r, column=col).value
        col += 1
        if col > 13:
            break
    return total_sum

def stdev(sheet_obj, r):
    deviation, col = 0, 2
    while(str(sheet_obj.cell(row=r, column=col).value).isdigit() == False):
        col += 1
    sum_value = annualSum(sheet_obj, r)
    while(sheet_obj.cell(row=r, column=col).value != None):
        deviation += (sheet_obj.cell(row=r, column=col).value - sum_value) ** 2
        col += 1
        if col > 13:
            break
    stdeviation = (deviation / (col - 2)) ** (1/2)
    return stdeviation

def XYZanal():
    path = r"C:\Users\19295\Downloads\project dataset.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    r = findRow(sheet_obj)
    xyzanlaysis, total = {}, 0
    while(sheet_obj.cell(row=r, column=1).value != None):
        code_item = sheet_obj.cell(row=r, column=1).value
        stdeviation = stdev(sheet_obj, r)
        total += stdeviation
        xyzanlaysis[code_item] = stdeviation
        r += 1

    xyzanlaysis = {i[0]: [i[1], i[1] / total * 100] for i in sorted(xyzanlaysis.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)}
    cumulative_percent, xCount, yCount, zCount = 0, 0, 0, 0

    for key in xyzanlaysis.keys():
        cumulative_percent = xyzanlaysis[key][1] + cumulative_percent
        xyzanlaysis[key].append(cumulative_percent)
        if cumulative_percent > 95:
            xyzanlaysis[key].append('Z')
            xCount += 1
        elif cumulative_percent > 80:
            xyzanlaysis[key].append('Y')
            yCount += 1
        else:
            xyzanlaysis[key].append('X')
            zCount += 1

    with st.container():
        col1, col2 = st.columns(2)
        with col1:
            labels = ['X', 'Y', 'Z']
            sizes = [xCount, yCount, zCount]
            fig, ax = plt.subplots()
            ax.pie(sizes, labels=labels, autopct='%1.1f%%')
            ax.axis('equal')
            st.pyplot(fig)

        with col2:
            barchart = pd.DataFrame(
                {'Count': [xCount, yCount, zCount]},
                index=['X', 'Y', 'Z']
            )
            st.bar_chart(barchart)

    XYZanalysis = pd.DataFrame(xyzanlaysis).T
    st.write(XYZanalysis.head(10))

XYZanal()
